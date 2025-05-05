import os, requests
import scrapy, time, openpyxl, random
from scrapy import signals
import pandas as pd
from openpyxl import load_workbook
from ..settings import USER_AGENT


class ExampleSpider(scrapy.Spider):
    name = "amazon"

    def __init__(self, url=None, category_name=None, filter_by=None, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.base_url = url if url else "https://default-url.com"
        self.category_name = category_name
        self.file_path = ''
        self.current_page = 1
        self.filter_by = filter_by
        self.scraped_urls = set()
        self.asins = set()
        self.data = []  # List to store all scraped offers
        self.results = []

    def start_requests(self):
        print(' ############################### self.base_url ############################### ', self.base_url)
        paginated_url = f"{self.base_url}&page={self.current_page}&ref=sr_nr_p_n_condition-type_1&s={self.filter_by}" # s=price-desc-rank (from highest price) s=popularity-rank
        headers = {"User-Agent": random.choice(USER_AGENT)}
        yield scrapy.Request(
            url=paginated_url,
            callback=self.parse,
            headers=headers
        )
        # or manualy random User-Agent logic:
        # headers = {"User-Agent": random.choice(self.settings.get("USER_AGENT"))}
        # yield scrapy.Request(url=paginated_url, callback=self.parse, headers=headers)



    def parse(self, response):
        products = response.css('div.s-main-slot div.s-result-item')
        new_products_found = False

        import re, urllib.parse

        def extract_asin(url):
            match = re.search(r'/dp/([A-Z0-9]+)|/gp/product/([A-Z0-9]+)', url)
            if match:
                return match.group(1) or match.group(2)  # Return the ASIN
            return ''


        count = 0
        asin = None

        for product in products:
            product_url = product.css('a.a-link-normal::attr(href)').get()

            if product_url: # and asin not in self.asins
                if product_url[:5] == "/sspa":
                    parsed_url = urllib.parse.parse_qs(urllib.parse.urlparse(product_url).query)
                    row_product_url = parsed_url.get("url", [""])[0]
                    product_url = urllib.parse.unquote(row_product_url)

                    asin = extract_asin(product_url)
                    if asin:
                        print(' ############################### asin ############################### ', asin)
                        # self.asins.add(asin)

                else:
                    asin = extract_asin(product_url)
                    
                self.asins.add(asin)
                new_products_found = True

                product_title = product.css('span::text').get()  # First attempt (may return 'Sponsorowane')
                # print(' ############################### product_title CLEAN ############################### ', product_title)

                is_sponsored = ''

                if product_title == "Sponsorowane":  
                    print(' ############################### Sponsorowane ? ############################### ', 'YES')
                    product_title = product.css('h2 span::text').get()  # Extract actual title only if 'Sponsorowane' a-size-base-plus
                    print(' ############################### Sponsorowane product_title ############################### ', product_title)
                    is_sponsored = "Sponsorowane"

                product_data = {
                    'date': time.strftime("%Y-%m-%d %H:%M:%S"),
                    'is_sponsored': is_sponsored,
                    "asin": asin, 
                    'title': product_title,
                    'image_url': product.css('img.s-image::attr(src)').get(),
                    'price': product.css('span.a-price-whole::text').get(),
                    'rating': product.css('span.a-icon-alt::text').get(),
                    'reviews': product.css('span.a-size-base::text').get(),  
                    'url': f"https://amazon.pl{product_url}",
                }

                count += 1
                print(" ############################### COUNT ############################### ", count, self.category_name)

                if product_data['title'] == "Sponsorowane" and count == 3:
                    self.logger.info(" ############################### THE END ############################### ")
                    self.save_to_excel()
                    raise scrapy.exceptions.CloseSpider("Skipping sponsored product.")

                self.data.append(product_data)  # Add product to list
                yield product_data


        if new_products_found:
            self.current_page += 1
            next_page_url = f"{self.base_url}&page={self.current_page}&ref=sr_nr_p_n_condition-type_1&s=popularity-rank"
            random_delay = random.uniform(1, 2) #random.uniform(3, 12) 
            yield scrapy.Request(url=next_page_url, callback=self.parse)

            if self.current_page % 10 == 0:
    
                long_random_delay = random.uniform(1, 2) # random.uniform(60, 120)
                time.sleep(long_random_delay)  # Longer delay every 20 pages
                self.save_to_excel()
                self.logger.info(f" ************* LONG DELAY {long_random_delay} seconds at {self.current_page} page ************* ")

                yield scrapy.Request(url=next_page_url, callback=self.parse)

            # if self.current_page == 20:
            #     self.logger.info("20 finish. Saving to Excel.")
            #     self.save_to_excel()  # Call the save function when scraping stops
            #     return
                
            else:
                time.sleep(random_delay)  # Sleep for a random time to avoid being blocked
                self.logger.info(f" ************* Scraping page {self.current_page}... ****** DELAY {random_delay} seconds ************* ")

                yield scrapy.Request(url=next_page_url, callback=self.parse)
        
        # else:
        #     self.logger.info("No new products found. Saving to Excel.")
        #     self.save_to_excel()  # Call the save function when scraping stops

        if self.current_page == 400:
            self.logger.info("Scrape limit 400. Saving to Excel.")
            self.save_to_excel()  # Call the save function when scraping stops




    def save_to_excel(self):

        DATA_DIR = f"amz/products/{self.category_name}/{time.strftime('%Y-%m-%d')}/"
        os.makedirs(DATA_DIR, exist_ok=True)

        filename = f"{self.category_name}.xlsx"

        file_path = os.path.join(DATA_DIR, filename)
        self.file_path = file_path

        # Convert data to a Pandas DataFrame
        df = pd.DataFrame(self.data)

        # Clean data
        df["price"] = pd.to_numeric(df["price"].str.replace("\xa0", "", regex=False), errors="coerce")
        df["rating"] = pd.to_numeric(df["rating"].str.extract(r'(\d+)')[0], errors="coerce")

        # Remove duplicate ASINs (keeping first occurrence)
        df = df.drop_duplicates(subset=["asin"], keep="first")

        sorted_df = df.sort_values(by="price", ascending=False)

        try:
            # Load existing workbook
            workbook = load_workbook(file_path)
            sheet = workbook.active

            # Read existing ASINs to avoid duplicates
            existing_asins = set()
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming headers exist
                existing_asins.add(row[0])  # Assuming 'asin' is the first column

            # Filter out ASINs that are already in the file
            sorted_df = sorted_df[~sorted_df["asin"].isin(existing_asins)]

            # Append only new unique rows
            for _, row in sorted_df.iterrows():
                sheet.append(row.tolist())

            # Save changes
            workbook.save(file_path)

            self.logger.info(f" ############################### Data saved to '{file_path}', only unique ASINs recorded. ###############################")
            self.data = []  # Clear the data list after saving

            return True

            

        except FileNotFoundError:
            # If file doesn't exist, create a new one with unique ASINs
            sorted_df.to_excel(file_path, index=False, engine="openpyxl")


class UrlsSpider(scrapy.Spider):
    name = "urls"

    def __init__(self, category=None, file_path=None, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # self.urls = urls.split(',')
        self.category = category
        self.file_path = file_path
        self.df = pd.read_excel(file_path)
        self.current_page = 0
        self.results = []
        self.polish_months = {
            "stycznia": "01", "lutego": "02", "marca": "03", "kwietnia": "04",
            "maja": "05", "czerwca": "06", "lipca": "07", "sierpnia": "08",
            "września": "09", "października": "10", "listopada": "11", "grudnia": "12"
        }




    def extract_first_availability_date(self, response):
        # Try from bullet-style product details (most common)
        value = response.xpath(
            '//th[contains(text(), "Data pierwszej dostępności")]/following-sibling::td/text()'
        ).get()
        
        if value:
            return value.strip()

        # Fallback 1: Try from div-based detail sections
        value = response.xpath(
            '//span[contains(text(), "Data pierwszej dostępności")]/following-sibling::span/text()'
        ).get()
        
        if value:
            return value.strip()
        
        # Fallback 2: Try any row that has the label in a div table
        value = response.xpath(
            '//td[contains(text(), "Data pierwszej dostępności")]/following-sibling::td/text()'
        ).get()

        if value:
            return value.strip()

        # Fallback 3: Check bullet section with different nesting
        # for row in response.css("#detailBullets_feature_div li"):
        #     label = row.css("span.a-text-bold::text").get()
        #     if label and "Data pierwszej dostępności" in label:
        #         date = row.css("span::text")[-1].get()
        #         return date.strip()
        import re
        def clean_text(text):
            if not text:
                return ""
            return re.sub(r"[\u200e\u200f\u202a-\u202e]", "", text).strip()
        for row in response.css("#detailBullets_feature_div li"):

            label = row.css("span.a-text-bold::text").get()
            if label and "Data pierwszej dostępności" in clean_text(label):
                # Get all span texts, the second one is usually the value
                spans = row.css("span::text").getall()
                if len(spans) > 1:
                    return clean_text(spans[-1])

        return None

    

    def start_requests(self):
        for url in self.df["url"].dropna():
            # Validate URL format before making a request
            if not isinstance(url, str) or not url.startswith(("http://", "https://")):
                self.logger.warning(f" ------------------------- Skipping invalid URL ------------------------- : {url}")
                continue  # Skip malformed URLs
            
            # time.sleep(random.uniform(3, 5))  # Random delay between requests
            # headers = {"User-Agent": random.choice(self.settings.get("USER_AGENT", []))}
            headers = {"User-Agent": random.choice(USER_AGENT)}

            yield scrapy.Request(url=url, callback=self.parse, headers=headers)

    def parse(self, response):

        import re, urllib.parse, datetime

        asin = None
        def extract_asin(url):
            match = re.search(r'/dp/([A-Z0-9]+)|/gp/product/([A-Z0-9]+)', url)
            if match:
                return match.group(1) or match.group(2)  # Return the ASIN
            return ''
        
        # first_availability_date = response.xpath(
        #     '//th[contains(text(), "Data pierwszej dostępności")]/following-sibling::td/text()'
        # ).get()
        # first_availability_date = response.css("#productDetails_detailBullets_sections1 td.a-size-base.prodDetAttrValue::text").get()
        first_availability_date = self.extract_first_availability_date(response) #response.css("#productDetails_detailBullets_sections1 tr:nth-child(3) td::text").get()
        print(' ++++++++++++++++++++++++++++++++++++++ first_availability_date ++++++++++++++++++++++++++++++++++++++ ', first_availability_date)
        # date = "---"
        # if first_availability_date is not None:
        #     day, month_name, year = first_availability_date.split()
        #     month = self.polish_months.get(month_name)
        #     formated_date = f"{year}-{month}-{day.zfill(2)}"
        #     date = formated_date.datetime.strptime(formated_date, "%y-%m-%d")
        print(' ++++++++++++++++++++++++++++++++++++++ response.url ++++++++++++++++++++++++++++++++++++++ ', response.url)
        asin = extract_asin(response.url)

        self.df.loc[self.df["asin"] == asin, "first_availability_date"] = first_availability_date

        self.current_page += 1

        # if self.current_page % 50 == 0:
        self.save_date_to_excel()


    def save_date_to_excel(self):
        print(' ++++++++++++++++++++++++++++++++++++++ save_date_to_excel calling ++++++++++++++++++++++++++++++++++++++ ',)
        try:
            self.df.to_excel(self.file_path, index=False)
            self.logger.info(f" ++++++++++++++++++++++++++++++++++++++ Updated Excel file saved to: {self.file_path} ++++++++++++++++++++++++++++++++++++++ ")
            self.current_page == 0
        except Exception as e:
            self.logger.error(f" ++++++++++++++++++++++++++++++++++++++ Failed to save Excel file: {e} ++++++++++++++++++++++++++++++++++++++ ")

